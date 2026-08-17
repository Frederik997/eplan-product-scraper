from playwright.sync_api import sync_playwright
from openpyxl import Workbook


BASE_URL = "https://dataportal.eplan.com"


INFORMATION = (
    "Part number",
    "Manufacturer",
    "Designation",
    "Description",
    "Product group",
    "Category",
)

page_number = 1

products = []


#Failed to process info due to connection error or something outside of our control
failed_products = []


with sync_playwright() as pw:

    browser = pw.chromium.launch(headless=True)

    page = browser.new_page()

    page.set_default_timeout(60000)

    page.goto(
        BASE_URL + "/parts/list?page=" + str(page_number) + "&perpage=50"
    )

    next_btn = page.locator(
        "eplan-icon-button[data-test='Pagination-NextPage']"
    )

    next_btn.wait_for()

    links = page.locator(
        "a[href^='/part-details/'][data-test]"
    )

    while True:

        # tell us which pagination page we're processing
        print("================================")
        print("Processing pagination page:", page_number)
        print("================================")

        links.nth(0).wait_for()

        hrefs = []

        for i in range(links.count()):

            hrefs.append(
                links.nth(i).get_attribute("href")
            )

        cleaned_hrefs = []

        for href in hrefs:

            cleaned_hrefs.append(
                BASE_URL + href.split("?")[0]
            )

        # Process every product on this pagination page
        for cleaned_href in cleaned_hrefs:

            print("Opening:", cleaned_href)

            try:

                page.goto(cleaned_href)
                table = page.locator("table").nth(0)


                table.wait_for()
                page.wait_for_timeout(1000)
                rows = table.locator(
                    "tbody > tr"
                )

                rows.nth(0).wait_for()
                product = {}

                for i in range(rows.count()):

                    row = rows.nth(i)

                    cells = row.locator("td")

                    label = cells.nth(0).inner_text().strip()

                    if label in INFORMATION:

                        value = cells.nth(1).inner_text().strip()

                        product[label] = value

                products.append(product)

                print(
                    "Scraped:",
                    product.get("Part number", "N/A")
                )

            except Exception as e:

                print("FAILED:", cleaned_href)
                print("Error:", e)


                failed_products.append({
                        "url": cleaned_href,
                        "page": page_number
                    })
                # Skip this product and continue
                continue

        print(
            "Finished pagination page:",
            page_number
        )

        print(
            "Total products scraped:",
            len(products)
        )

        # Return to the list page
        page.goto(
          BASE_URL + "/parts/list?page=" + str(page_number) + "&perpage=50"
        )

        links.nth(0).wait_for()

        # Check if this is the last page
        if next_btn.get_attribute("disabled") is not None:

            print("Reached the last page.")
            break

        # Move to next page
        page_number += 1
        print(products)
        print(
            "Moving to pagination page:",
            page_number
        )

        page.goto(
            BASE_URL + "/parts/list?page=" + str(page_number) + "&perpage=50"
        )

        links.nth(0).wait_for()

    browser.close()


# Sort by manufacturer
products.sort(
    key=lambda product: product.get("Manufacturer", "")
)


# Create Excel workbook
workbook = Workbook()

sheet = workbook.active


headers = [
    "Part number",
    "Manufacturer",
    "Designation",
    "Description",
    "Product group"
]


# Headers
for column, header in enumerate(headers, start=1):

    sheet.cell(
        row=1,
        column=column,
        value=header
    )


# Products
for row, product in enumerate(products, start=2):

    for column, header in enumerate(headers, start=1):

        value = product.get(header, "")

        if value == "":
            value = "N/A"

        sheet.cell(
            row=row,
            column=column,
            value=value
        )


# Save Excel
workbook.save("products.xlsx")

print("Finished!")
print("Total products:", len(products))